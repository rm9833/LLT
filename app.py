import streamlit as st
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color, GradientFill
from googletrans import Translator, LANGUAGES
import os
from pathlib import Path
import tempfile
import time
import os

os.makedirs("Downloads", exist_ok=True)

# Title
st.title("Excel Translator App")

global cellStyle

cellStyle = []

translator = Translator()

# Step 1: Upload file
uploaded_file = st.file_uploader("Upload a Excel file", type=["xlsx"])

def open_workbook(path):
    workbook = load_workbook(filename=path)
    # get_chart_details(workbook)
    sheetValueArray = []
    for sheet in workbook.sheetnames:
        sheetData = workbook[sheet]
        for column in sheetData[sheetData.calculate_dimension()]:
            cellArray = []
            for cell in column:
                # print(cell.fill)
                cellname = (f"{cell.column_letter}{cell.row}")
                try:
                    if cell.fill.patternType:
                        styleType = {
                            'cellName': cellname,
                            'type': 'PatternFill',
                        }
                        cell_format = {
                            'sheet' : {
                                'sheetname' : sheet  
                            },
                            'cellname' : {
                                'position': cellname,
                                'value': cell.value
                            },
                            'font': {
                                'name': cell.font.name,
                                'sz': cell.font.sz,
                                'b': cell.font.b,
                                'i': cell.font.i,
                                'charset': cell.font.charset,
                                'u': cell.font.u,
                                'strike': cell.font.strike,
                                'color': {
                                            'rgb': cell.font.color.rgb if cell.font.color.type == 'rgb' else "00000000",
                                            'indexed': cell.font.color.indexed if cell.font.color.type == "Indexed" else None,
                                            'auto': cell.font.color.auto if not(cell.font.color.auto) == None else False,
                                            'theme': cell.font.color.theme if cell.font.color.type == 'theme' else None,
                                            'tint' : cell.font.color.tint,
                                            'index' : cell.font.color.index,
                                            'type' : cell.font.color.type
                                        } if not(cell.font.color) == None else None,
                                'scheme': cell.font.scheme,
                                'family': cell.font.family,
                                'size': cell.font.size,
                                'bold': cell.font.bold,
                                'italic': cell.font.italic,
                                'strikethrough': cell.font.strikethrough,
                                'underline': cell.font.underline,
                                'vertAlign': cell.font.vertAlign,
                                'outline': cell.font.outline,
                                'shadow': cell.font.shadow,
                                'condense': cell.font.condense,
                                'extend': cell.font.extend
                            },
                            'fill': 
                                {
                                'patternType': cell.fill.patternType,
                                'fgColor': {
                                                'rgb': cell.fill.fgColor.rgb if cell.fill.fgColor.type == 'rgb' else "00000000",
                                                'indexed': cell.fill.fgColor.indexed if cell.fill.fgColor.type == "indexed" else None,
                                                'auto': cell.fill.fgColor.auto if not(cell.fill.fgColor.auto) == None else False,
                                                'theme': cell.fill.fgColor.theme if cell.fill.fgColor.type == 'theme' else None,
                                                'tint' : cell.fill.fgColor.tint,
                                                'index' : cell.fill.fgColor.index,
                                                'type' : cell.fill.fgColor.type
                                                },
                                'bgColor': {
                                                'rgb': cell.fill.bgColor.rgb if cell.fill.bgColor.type == 'rgb' else "00000000",
                                                'indexed': cell.fill.bgColor.indexed if cell.fill.bgColor.type == "indexed" else None,
                                                'auto': cell.fill.bgColor.auto if not(cell.fill.bgColor.auto) == None else False,
                                                'theme': cell.fill.bgColor.theme if cell.fill.bgColor.type == 'theme' else None,
                                                'tint' : cell.fill.bgColor.tint,
                                                'index' : cell.fill.bgColor.index,
                                                'type' : cell.fill.bgColor.type
                                                },
                                'fill_type': cell.fill.fill_type,
                                'start_color': {
                                                'rgb': cell.fill.start_color.rgb if cell.fill.start_color.type == 'rgb' else "00000000",
                                                'indexed': cell.fill.start_color.indexed if cell.fill.start_color.type == "indexed" else None,
                                                'auto': cell.fill.start_color.auto if not(cell.fill.start_color.auto) == None else False,
                                                'theme': cell.fill.start_color.theme if cell.fill.start_color.type == 'theme' else None,
                                                'tint' : cell.fill.start_color.tint,
                                                'index' : cell.fill.start_color.index,
                                                'type' : cell.fill.start_color.type
                                                },
                                'end_color': {
                                                'rgb': cell.fill.end_color.rgb if cell.fill.end_color.type == 'rgb' else "00000000",
                                                'indexed': cell.fill.end_color.indexed if cell.fill.end_color.type == "indexed" else None,
                                                'auto': cell.fill.end_color.auto if not(cell.fill.end_color.auto) == None else False,
                                                'theme': cell.fill.end_color.theme if cell.fill.end_color.type == 'theme' else None,
                                                'tint' : cell.fill.end_color.tint,
                                                'index' : cell.fill.end_color.index,
                                                'type' : cell.fill.end_color.type
                                                }  
                                },
                            'alignment': {
                                'horizontal': cell.alignment.horizontal,
                                'vertical': cell.alignment.vertical,
                                'textRotation': cell.alignment.textRotation,
                                'wrapText': cell.alignment.wrapText,
                                'shrinkToFit': cell.alignment.shrinkToFit,
                                'indent': cell.alignment.indent,
                                'relativeIndent': cell.alignment.relativeIndent,
                                'justifyLastLine': cell.alignment.justifyLastLine,
                                'readingOrder': cell.alignment.readingOrder,
                                'text_rotation': cell.alignment.text_rotation,
                                'wrap_text': cell.alignment.wrap_text,
                                'shrink_to_fit': cell.alignment.shrink_to_fit,
                                # 'mergeCell': cell.alignment.mergeCell if (isinstance(cell, MergedCell)) else None
                            },
                            'border': {
                                'left': cell.border.left.style,
                                'right': cell.border.right.style,
                                'top': cell.border.top.style,
                                'bottom': cell.border.bottom.style,
                                'diagonal': cell.border.diagonal.style,
                                # 'diagonal_direction': cell.border.diagonal_direction.style,
                                'vertical': cell.border.vertical,
                                'horizontal': cell.border.horizontal,
                                'diagonalUp': cell.border.diagonalUp,
                                'diagonalDown': cell.border.diagonalDown,
                                'outline': cell.border.outline,
                                'start': cell.border.start,
                                'end': cell.border.end,
                                'color': {
                                    'left': cell.border.left.color,
                                    'right': cell.border.right.color,
                                    'top': cell.border.top.color,
                                    'bottom': cell.border.bottom.color,
                                    'diagonal': cell.border.diagonal.color
                                    # 'diagonal_direction': cell.border.diagonal_direction.color,
                                }
                            }
                        }
                    
                    else:
                        styleType = {
                            'cellName': cellname,
                            'type': 'PatternFill',
                        }
                        cell_format = {
                            'sheet' : {
                                'sheetname' : sheet  
                            },
                            'cellname' : {
                                'position': cellname,
                                'value': cell.value
                            },
                            'font': {
                                'name': cell.font.name,
                                'sz': cell.font.sz,
                                'b': cell.font.b,
                                'i': cell.font.i,
                                'charset': cell.font.charset,
                                'u': cell.font.u,
                                'strike': cell.font.strike,
                                'color': {
                                            'rgb': cell.font.color.rgb if cell.font.color.type == 'rgb' else "00000000",
                                            'indexed': cell.font.color.indexed if cell.font.color.type == "Indexed" else None,
                                            'auto': cell.font.color.auto if not(cell.font.color.auto) == None else False,
                                            'theme': cell.font.color.theme if cell.font.color.type == 'theme' else None,
                                            'tint' : cell.font.color.tint,
                                            'index' : cell.font.color.index,
                                            'type' : cell.font.color.type
                                        } if not(cell.font.color) == None else None,
                                'scheme': cell.font.scheme,
                                'family': cell.font.family,
                                'size': cell.font.size,
                                'bold': cell.font.bold,
                                'italic': cell.font.italic,
                                'strikethrough': cell.font.strikethrough,
                                'underline': cell.font.underline,
                                'vertAlign': cell.font.vertAlign,
                                'outline': cell.font.outline,
                                'shadow': cell.font.shadow,
                                'condense': cell.font.condense,
                                'extend': cell.font.extend
                            },
                            'fill': 
                                {
                                'patternType': cell.fill.patternType,
                                'fgColor': {
                                                'rgb': cell.fill.fgColor.rgb if cell.fill.fgColor.type == 'rgb' else "00000000",
                                                'indexed': cell.fill.fgColor.indexed if cell.fill.fgColor.type == "indexed" else None,
                                                'auto': cell.fill.fgColor.auto if not(cell.fill.fgColor.auto) == None else False,
                                                'theme': cell.fill.fgColor.theme if cell.fill.fgColor.type == 'theme' else None,
                                                'tint' : cell.fill.fgColor.tint,
                                                'index' : cell.fill.fgColor.index,
                                                'type' : cell.fill.fgColor.type
                                                },
                                'bgColor': {
                                                'rgb': cell.fill.bgColor.rgb if cell.fill.bgColor.type == 'rgb' else "00000000",
                                                'indexed': cell.fill.bgColor.indexed if cell.fill.bgColor.type == "indexed" else None,
                                                'auto': cell.fill.bgColor.auto if not(cell.fill.bgColor.auto) == None else False,
                                                'theme': cell.fill.bgColor.theme if cell.fill.bgColor.type == 'theme' else None,
                                                'tint' : cell.fill.bgColor.tint,
                                                'index' : cell.fill.bgColor.index,
                                                'type' : cell.fill.bgColor.type
                                                },
                                'fill_type': cell.fill.fill_type,
                                'start_color': {
                                                'rgb': cell.fill.start_color.rgb if cell.fill.start_color.type == 'rgb' else "00000000",
                                                'indexed': cell.fill.start_color.indexed if cell.fill.start_color.type == "indexed" else None,
                                                'auto': cell.fill.start_color.auto if not(cell.fill.start_color.auto) == None else False,
                                                'theme': cell.fill.start_color.theme if cell.fill.start_color.type == 'theme' else None,
                                                'tint' : cell.fill.start_color.tint,
                                                'index' : cell.fill.start_color.index,
                                                'type' : cell.fill.start_color.type
                                                },
                                'end_color': {
                                                'rgb': cell.fill.end_color.rgb if cell.fill.end_color.type == 'rgb' else "00000000",
                                                'indexed': cell.fill.end_color.indexed if cell.fill.end_color.type == "indexed" else None,
                                                'auto': cell.fill.end_color.auto if not(cell.fill.end_color.auto) == None else False,
                                                'theme': cell.fill.end_color.theme if cell.fill.end_color.type == 'theme' else None,
                                                'tint' : cell.fill.end_color.tint,
                                                'index' : cell.fill.end_color.index,
                                                'type' : cell.fill.end_color.type
                                                }  
                                },
                            'alignment': {
                                'horizontal': cell.alignment.horizontal,
                                'vertical': cell.alignment.vertical,
                                'textRotation': cell.alignment.textRotation,
                                'wrapText': cell.alignment.wrapText,
                                'shrinkToFit': cell.alignment.shrinkToFit,
                                'indent': cell.alignment.indent,
                                'relativeIndent': cell.alignment.relativeIndent,
                                'justifyLastLine': cell.alignment.justifyLastLine,
                                'readingOrder': cell.alignment.readingOrder,
                                'text_rotation': cell.alignment.text_rotation,
                                'wrap_text': cell.alignment.wrap_text,
                                'shrink_to_fit': cell.alignment.shrink_to_fit,
                                # 'mergeCell': cell.alignment.mergeCell if (isinstance(cell, MergedCell)) else None
                            },
                            'border': {
                                'left': cell.border.left.style,
                                'right': cell.border.right.style,
                                'top': cell.border.top.style,
                                'bottom': cell.border.bottom.style,
                                'diagonal': cell.border.diagonal.style,
                                # 'diagonal_direction': cell.border.diagonal_direction.style,
                                'vertical': cell.border.vertical,
                                'horizontal': cell.border.horizontal,
                                'diagonalUp': cell.border.diagonalUp,
                                'diagonalDown': cell.border.diagonalDown,
                                'outline': cell.border.outline,
                                'start': cell.border.start,
                                'end': cell.border.end,
                                'color': {
                                    'left': cell.border.left.color,
                                    'right': cell.border.right.color,
                                    'top': cell.border.top.color,
                                    'bottom': cell.border.bottom.color,
                                    'diagonal': cell.border.diagonal.color
                                    # 'diagonal_direction': cell.border.diagonal_direction.color,
                                }
                            }
                        }
                    cellStyle.append(styleType)     
                    cellArray.append(cell_format) 
                except AttributeError:
                    if cell.fill.type:
                        styleType = {
                            'cellName': cellname,
                            'type': 'GradientFill',
                        }
                        cell_format = {
                            'sheet' : {
                                'sheetname' : sheet  
                            },
                            'cellname' : {
                                'position': cellname,
                                'value': cell.value
                            },
                            'font': {
                                'name': cell.font.name,
                                'sz': cell.font.sz,
                                'b': cell.font.b,
                                'i': cell.font.i,
                                'charset': cell.font.charset,
                                'u': cell.font.u,
                                'strike': cell.font.strike,
                                'color': {
                                            'rgb': cell.font.color.rgb if cell.font.color.type == 'rgb' else "00000000",
                                            'indexed': cell.font.color.indexed if cell.font.color.type == "Indexed" else None,
                                            'auto': cell.font.color.auto if not(cell.font.color.auto) == None else False,
                                            'theme': cell.font.color.theme if cell.font.color.type == 'theme' else None,
                                            'tint' : cell.font.color.tint,
                                            'index' : cell.font.color.index,
                                            'type' : cell.font.color.type
                                        } if not(cell.font.color) == None else None,
                                'scheme': cell.font.scheme,
                                'family': cell.font.family,
                                'size': cell.font.size,
                                'bold': cell.font.bold,
                                'italic': cell.font.italic,
                                'strikethrough': cell.font.strikethrough,
                                'underline': cell.font.underline,
                                'vertAlign': cell.font.vertAlign,
                                'outline': cell.font.outline,
                                'shadow': cell.font.shadow,
                                'condense': cell.font.condense,
                                'extend': cell.font.extend
                            },
                            'GradientFill':
                            {
                                'type': cell.fill.type,
                                'degree': cell.fill.degree,
                                'left': cell.fill.left,
                                'right': cell.fill.right,
                                'top': cell.fill.top,
                                'bottom': cell.fill.bottom,
                                'stop': cell.fill.stop,
                            },
                            'alignment': {
                                'horizontal': cell.alignment.horizontal,
                                'vertical': cell.alignment.vertical,
                                'textRotation': cell.alignment.textRotation,
                                'wrapText': cell.alignment.wrapText,
                                'shrinkToFit': cell.alignment.shrinkToFit,
                                'indent': cell.alignment.indent,
                                'relativeIndent': cell.alignment.relativeIndent,
                                'justifyLastLine': cell.alignment.justifyLastLine,
                                'readingOrder': cell.alignment.readingOrder,
                                'text_rotation': cell.alignment.text_rotation,
                                'wrap_text': cell.alignment.wrap_text,
                                'shrink_to_fit': cell.alignment.shrink_to_fit,
                                # 'mergeCell': cell.alignment.mergeCell if (isinstance(cell, MergedCell)) else None
                            },
                            'border': {
                                'left': cell.border.left.style,
                                'right': cell.border.right.style,
                                'top': cell.border.top.style,
                                'bottom': cell.border.bottom.style,
                                'diagonal': cell.border.diagonal.style,
                                # 'diagonal_direction': cell.border.diagonal_direction.style,
                                'vertical': cell.border.vertical,
                                'horizontal': cell.border.horizontal,
                                'diagonalUp': cell.border.diagonalUp,
                                'diagonalDown': cell.border.diagonalDown,
                                'outline': cell.border.outline,
                                'start': cell.border.start,
                                'end': cell.border.end,
                                'color': {
                                    'left': cell.border.left.color,
                                    'right': cell.border.right.color,
                                    'top': cell.border.top.color,
                                    'bottom': cell.border.bottom.color,
                                    'diagonal': cell.border.diagonal.color
                                    # 'diagonal_direction': cell.border.diagonal_direction.color,
                                }
                            }
                        } 
                        cellStyle.append(styleType)     
                        cellArray.append(cell_format)
            sheetValueArray.append(cellArray)
        
        # print(sheetData._charts)
        
    return sheetValueArray, cellStyle
        
def translate_cell_values(data, language):
    for row in data:
        for cell in row:
            value = cell['cellname']['value']
            if not(value) == "":
                if not(value) == None and type(value) == str:
                    translatedText = translate_text(value, language)
                    
                    cell['cellname']['value'] = translatedText
            
def translate_text(user_input,user_lang):
    try:
        translation = translator.translate(text = user_input, dest=user_lang)
        return (f"{translation.text}")
    except Exception as e:
        return user_input
    # print(f"{translation.origin} ({translation.src}) --> {translation.text} ({translation.dest})")

def set_cell_format(cell, font, fill, alignment, border, fillType):
    cell.font = Font(
        name=font['name'],
        sz=font['sz'],
        b=font['b'],
        i=font['i'],
        charset=font['charset'],
        u=font['u'],
        strike=font['strike'],
        color = ((Color(rgb=font['color']['rgb'], tint= font['color']['tint'])) if font['color']['type'] == 'rgb' 
                else ((Color(theme=font['color']['theme'], tint= font['color']['tint'])) if font['color']['type'] == 'theme'
                else (Color(index=font['color']['index'],  tint= font['color']['tint']))))
                if not(font['color']) == None else None,
        scheme=font['scheme'],
        family=font['family'],
        size=font['size'],
        bold=font['bold'],
        italic=font['italic'],
        strikethrough=font['strikethrough'],
        underline=font['underline'],
        vertAlign=font['vertAlign'],
        outline=font['outline'],
        shadow=font['shadow'],
        condense=font['condense'],
        extend=font['extend']
    )
    cell.fill = PatternFill(
        patternType=fill['patternType'],
        fgColor=((Color(rgb=fill['fgColor']['rgb'], tint= fill['fgColor']['tint'])) if fill['fgColor']['type'] == 'rgb' 
                else ((Color(theme=fill['fgColor']['theme'], tint= fill['fgColor']['tint'])) if fill['fgColor']['type'] == 'theme'
                else (Color(index=fill['fgColor']['index'],  tint= fill['fgColor']['tint']))))
                if not(fill['fgColor']) == None else None,
        bgColor=((Color(rgb=fill['bgColor']['rgb'], tint= fill['bgColor']['tint'])) if fill['bgColor']['type'] == 'rgb' 
                else ((Color(theme=fill['bgColor']['theme'], tint= fill['bgColor']['tint'])) if fill['bgColor']['type'] == 'theme'
                else (Color(index=fill['bgColor']['index'],  tint= fill['bgColor']['tint']))))
                if not(fill['bgColor']) == None else None,
        fill_type=fill['fill_type'],
        start_color=((Color(rgb=fill['start_color']['rgb'], tint= fill['start_color']['tint'])) if fill['start_color']['type'] == 'rgb' 
                else ((Color(theme=fill['start_color']['theme'], tint= fill['start_color']['tint'])) if fill['start_color']['type'] == 'theme'
                else (Color(index=fill['start_color']['index'],  tint= fill['start_color']['tint']))))
                if not(fill['start_color']) == None else None,
        end_color=((Color(rgb=fill['end_color']['rgb'], tint= fill['end_color']['tint'])) if fill['end_color']['type'] == 'rgb' 
                else ((Color(theme=fill['end_color']['theme'], tint= fill['end_color']['tint'])) if fill['end_color']['type'] == 'theme'
                else (Color(index=fill['end_color']['index'],  tint= fill['end_color']['tint']))))
                if not(fill['end_color']) == None else None,
    ) if fillType == 'PatternFill' else GradientFill(
        type=fill['type'],
        degree=fill['degree'],
        left=fill['left'],
        right=fill['right'],
        top=fill['top'],
        bottom=fill['bottom'],
        stop=fill['stop']
    )
    cell.alignment = Alignment(
        horizontal=alignment['horizontal'],
        vertical=alignment['vertical'],
        textRotation=alignment['textRotation'],
        wrapText=alignment['wrapText'],
        shrinkToFit=alignment['shrinkToFit'],
        indent=alignment['indent'],
        relativeIndent=alignment['relativeIndent'],
        justifyLastLine=alignment['justifyLastLine'],
        readingOrder=alignment['readingOrder'],
        text_rotation=alignment['text_rotation'],
        wrap_text=alignment['wrap_text'],
        shrink_to_fit=alignment['shrink_to_fit'],
        # mergeCell=alignment['mergeCell']
    )
    cell.border = Border(
        left=Side(style=border['left'], color=border['color']['left']),
        right=Side(style=border['right'], color=border['color']['right']),
        top=Side(style=border['top'], color=border['color']['top']),
        bottom=Side(style=border['bottom'], color=border['color']['bottom']),
        diagonal=Side(style=border['diagonal'], color=border['color']['diagonal']),
        # diagonal_direction: Any | None = None,
        vertical= border['vertical'],
        horizontal= border['horizontal'],
        diagonalUp= border['diagonalUp'],
        diagonalDown= border['diagonalDown'],
        outline= border['outline'],
        start= border['start'],
        end= border['end']
    )
    
         
def createTranslateExcel(data, language, cellStyle, orginal_filename):
    workbook = openpyxl.Workbook()
    i = 0

    for row in data:
        for cell_data in row:
            sheetname = cell_data['sheet']['sheetname']
            if sheetname not in workbook.sheetnames:
                workbook.create_sheet(title=sheetname)
            sheet = workbook[sheetname]

            fillType = cellStyle[i]['type']
            i += 1
            position = cell_data['cellname']['position']
            value = cell_data['cellname']['value']
            font = cell_data['font']
            alignment = cell_data['alignment']
            border = cell_data['border']

            cell = sheet[position]
            cell.value = value
            if fillType == 'PatternFill':
                set_cell_format(cell, font, cell_data['fill'], alignment, border, fillType)
            elif fillType == 'GradientFill':
                set_cell_format(cell, font, cell_data['GradientFill'], alignment, border, fillType)

    # Remove the default sheet created by openpyxl if it is empty
    if 'Sheet' in workbook.sheetnames and workbook['Sheet'].max_row == 1 and workbook['Sheet'].max_column == 1:
        workbook.remove(workbook['Sheet'])

    # Construct the file path to the Downloads folder
    downloads_path = "Downloads"
    file_path = os.path.join(downloads_path, language + "_" + orginal_filename)

    # Save the workbook
    workbook.save(file_path)
    return file_path
    


# Step 2: Language selection (after file upload)
if uploaded_file is not None:
    
    languages = {}

    for code, language in LANGUAGES.items():
        languages[language.title()] = code
        
    lang_name = st.selectbox("Select target language", list(languages.keys()))
    lang_code = languages[lang_name]

    # Step 3: Submit button
    if st.button("Translate"):
        progress_bar  = st.progress(0, text="Processing...")
        # Read the file content
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(uploaded_file.getbuffer())
            uploaded_file_path = tmp_file.name
        
        for i in range(1, 20):
            progress_bar.progress(i, text=f"Processing... {i}%")
            time.sleep(0.1)
            
        
        filename_with_extension = os.path.basename(uploaded_file_path)
        workbookValueArray, cellStyle = open_workbook(uploaded_file_path)
        for i in range(20, 40):
            progress_bar.progress(i, text=f"Processing... {i}%")
            time.sleep(0.1)
        translate_cell_values(workbookValueArray, lang_code)
        for i in range(40, 80):
            progress_bar.progress(i, text=f"Processing... {i}%")
            time.sleep(0.1)
        tmp_path = createTranslateExcel(workbookValueArray, lang_code, cellStyle, filename_with_extension)
        st.session_state.tmp_path = tmp_path
        for i in range(80, 100):
            progress_bar.progress(i, text=f"Processing... {i}%")
            time.sleep(0.1)
        # Step 4: Download button
        progress_bar.empty()
        st.session_state.file_processed = True
        if st.session_state.file_processed and st.session_state.tmp_path:
            st.success("Translation completed!")
            st.write("Click the button below to download the translated file.")
            col1, col2 = st.columns([2, 1])
            with open(st.session_state.tmp_path, "rb") as f:
                col1.download_button(
                    label="Download Translated File",
                    data=f,
                    file_name=f"translated_{lang_code}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                
            if col2.button("Reset"):
                os.remove(st.session_state.tmp_path)
                st.session_state.file_processed = False
                st.session_state.tmp_path = None
                st.experimental_rerun()
        else:
            st.error("Error: File not processed or path not set.")

